from capedata import MergeShortNameData
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
sys.path.append('.\test_SI')
#from test_SI import compute_and_plot
#from test_SI import test_and_plot_orig
from metric_names import MetricName
from metric_names import NonMetricName
from sw_bias import compute_sw_bias
import openpyxl
import openpyxl.chart
import pandas as pd
import numpy as np  
import sys
import importlib
import re
from pathlib import Path
from metric_names import KEY_METRICS
from generate_SI import compute_only, mk_data as mk_si_data
from generate_SI import BASIC_NODE_SET
from generate_SI import ALL_NODE_SET
from generate_SI import SiData, StaticSiPlot
from capeplot import CapacityData
from capelib import crossjoin
import os
from argparse import ArgumentParser
import gc

class SatAnalysisSettings:
  SW_BIAS_IP = [MetricName.COUNT_OPS_VEC_PCT, MetricName.COUNT_VEC_TYPE_OPS_PCT, MetricName.COUNT_OPS_FMA_PCT,
                MetricName.COUNT_OPS_CVT_PCT, MetricName.COUNT_OPS_DIV_PCT,
                MetricName.SRC_RHS_OP_COUNT, MetricName.SRC_RECURRENCE_B, MetricName.SRC_CLU_SCORE]
  SW_BIAS_COLUMNS = ['Nd_CNVT_OPS', 'Nd_VEC_OPS', 'Nd_DIV_OPS', 'Nd_FMA_OPS', 'Nd_ISA_EXT_TYPE', 'Nd_clu_score', 'Nd_Recurrence' , 'Nd_RHS',
                     'Neg_SW_Bias', 'Pos_SW_Bias', 'Net_SW_Bias', NonMetricName.SI_TIER_NORMALIZED]
  ALL_CU_NODES = [ MetricName.STALL_SB_PCT, MetricName.STALL_LM_PCT, MetricName.STALL_LB_PCT, MetricName.STALL_FE_PCT, MetricName.STALL_RS_PCT]
  SELECTED_CU_NODES = [ MetricName.STALL_SB_PCT, MetricName.STALL_LM_PCT, MetricName.STALL_LB_PCT]
  MEM_NODES = [MetricName.RATE_L1_GB_P_S, MetricName.RATE_L2_GB_P_S, MetricName.RATE_L3_GB_P_S, MetricName.RATE_RAM_GB_P_S]
  FP_REG_NODES = [MetricName.RATE_FP_GFLOP_P_S, MetricName.RATE_REG_SIMD_GB_P_S]
  ALL_NODE_LIST =  MEM_NODES + FP_REG_NODES + [MetricName.STALL_SB_PCT, MetricName.STALL_LM_PCT, MetricName.STALL_LB_PCT]
  BASIC_INPUTS =  ALL_NODE_LIST + [MetricName.STALL_FE_PCT, MetricName.STALL_RS_PCT]
  BASIC_OUTPUTS = [NonMetricName.SI_CLUSTER_NAME, NonMetricName.SI_SAT_NODES, NonMetricName.SI_SAT_TIER] 
  CU_NODE_SET={MetricName.STALL_FE_PCT, MetricName.STALL_LB_PCT, MetricName.STALL_SB_PCT, MetricName.STALL_LM_PCT, MetricName.STALL_RS_PCT}
  CU_NODE_DICT={MetricName.STALL_FE_PCT:'FE [GW/s]', MetricName.STALL_LB_PCT:'LB [GW/s]', MetricName.STALL_SB_PCT:'SB [GW/s]', MetricName.STALL_LM_PCT:'LM [GW/s]', MetricName.STALL_RS_PCT:'RS [GW/s]'}
  BASIC_NODE_LIST=list(BASIC_NODE_SET)
  CU_SAT_THRESHOLD_DEFAULT = 0.25
  SAT_THRESHOLD_DEFAULT = 0.1

  def __init__(self, cuSatThreshold = CU_SAT_THRESHOLD_DEFAULT, satThreshold = SAT_THRESHOLD_DEFAULT, disable=False, 
               cu_traffic = [ MetricName.STALL_SB_PCT, MetricName.STALL_LM_PCT, MetricName.STALL_LB_PCT], 
               compute_stats = True, uniform_gap = False, run_sw_bias=True, chosen_node_set=ALL_NODE_SET):
    self.cuSatThreshold = cuSatThreshold
    self.satThreshold = satThreshold
    self.disable = disable
    self.percentsToCheck = cu_traffic
    self.compute_stats = compute_stats
    self.compute_stats = True
    self.fp_reg_rate_metrics = self.FP_REG_NODES
    self.mem_traffic = self.MEM_NODES
    self.non_cu_tiering_metrics = self.fp_reg_rate_metrics + self.mem_traffic
    self.non_traffic_tiering_metrics = self.fp_reg_rate_metrics + self.percentsToCheck
    self.tiering_metrics = self.mem_traffic + self.non_traffic_tiering_metrics
    self.uniform_gap = uniform_gap
    self.run_sw_bias = run_sw_bias
    self.chosen_node_set = chosen_node_set

  def set_chosen_node_set(self, chosen_node_set):
    self.chosen_node_set = chosen_node_set

  def input_columns(self):
    inputs =  self.BASIC_INPUTS + SiData.capacities(self.chosen_node_set)
    if self.run_sw_bias:
      inputs = inputs + self.SW_BIAS_IP
    return inputs

  def output_columns(self):
    outputs = self.BASIC_OUTPUTS
    if self.run_sw_bias:
      outputs = outputs + self.SW_BIAS_COLUMNS
    return outputs
      
    

# percent within max in column for color
# TODO unused at the moment
#@satThreshold = 0.10
#cuSatThreshold = 0.25

DO_SUB_CLUSTERING = False
DO_DEBUG_LOGS = False
PRINT_ALL_CLUSTERS = False
PRINT_COLOURED_TIERS = False
RUN_SI = True


# This has to be identical to BASIC_NODE_SET in generate_SI
#BASIC_NODE_LIST=['L1 [GB/s]', 'L2 [GB/s]', 'L3 [GB/s]', 'FLOP [GFlop/s]', 'VR [GB/s]', 'RAM [GB/s]']

#ALL_NODE_LIST =  [ "register_simd_rate_gb/s", "flop_rate_gflop/s", "l1_rate_gb/s", "l2_rate_gb/s", "l3_rate_gb/s", "ram_rate_gb/s",
#                   "%sb", "%lm", "%rs", "%lb"]

# arith
subNodeTrafficToCheck = [ MetricName.STALL_SB_PCT, MetricName.STALL_LM_PCT, MetricName.STALL_LB_PCT, MetricName.STALL_FE_PCT, MetricName.STALL_RS_PCT]

# OUTPUT_COLUMNS=[NonMetricName.SI_CLUSTER_NAME, NonMetricName.SI_SAT_NODES, NonMetricName.SI_SAT_TIER] + SW_BIAS_COLUMNS
# NEEDED_CLUSTER_DF_COLUMNS = KEY_METRICS+ OUTPUT_COLUMNS + ALL_NODE_LIST + [MetricName.STALL_FE_PCT, MetricName.STALL_RS_PCT]
# if RUN_SW_BIAS:
#   NEEDED_CLUSTER_DF_COLUMNS = NEEDED_CLUSTER_DF_COLUMNS + SW_BIAS_IP
# NEEDED_TEST_DF_COLUMNS = NEEDED_CLUSTER_DF_COLUMNS


# Commented out the following global variables as it looks like something we can compute locally using Pandas operations.
# # this dict contains columns + rows of those columns that need to be colored
# # assumption is that you don't add any more rows else dict becomes out of dat
# maxOfColumn = {}
# coloured_maxOfColumn = {}

# # Initializes a list of things in a dict as lists
# def initDict(maxOfColumn, listOfColumns):
#   for i in listOfColumns:
#     maxOfColumn[i] = []

# # init lists
# initDict(maxOfColumn, trafficToCheck)
# #initDict(maxOfColumn, percentsToCheck)
# initDict(maxOfColumn, archIntensityToCheck)

def compute_tier_gap(settings, satSetDF, traffic, cu_traffic):
  traffic_max = satSetDF[traffic].max() 
  cu_max = satSetDF[cu_traffic].max()
  traffic_gap = (traffic_max * settings.satThreshold).clip(lower=0)
  cu_gap = (cu_max * settings.cuSatThreshold).clip(lower=0)
  return [traffic_gap, cu_gap]

def compute_node_thresholds(settings, satSetDF, traffic, cu_traffic, uniform_gap):
  traffic_max = satSetDF[traffic].max() 
  cu_max = satSetDF[cu_traffic].max()
  if uniform_gap is None:
    traffic_gap, cu_gap = compute_tier_gap(settings, satSetDF, traffic, cu_traffic)
  else:
    traffic_gap, cu_gap = uniform_gap

  trafficThresholds = (traffic_max - traffic_gap).clip(lower=0)
  cuThresholds = cu_max.copy()
  cuSmallMask = cuThresholds < 50
  cuThresholds[cuSmallMask] = np.nan
  cuThresholds[~cuSmallMask] = (cuThresholds[~cuSmallMask] - cu_gap[~cuSmallMask]).clip(lower=0)
  # For threshold lower than 50, set it to infinity so no codelets can saturate.
  return pd.concat([trafficThresholds , cuThresholds], axis=0)
  

def satTrafficListToSiSatNodes(settings, df):
  return df.apply(lambda x: set(settings.BASIC_NODE_LIST) | {settings.CU_NODE_DICT[n] for n in set(x['satTrafficList']) & settings.CU_NODE_SET}, axis=1)
  
def compute_cluster_names(settings, tiers_sat_nodes_mask, sat_nodes):
  tiered_mask = ~tiers_sat_nodes_mask['Tier'].isna()
  sat_nodes_only=tiers_sat_nodes_mask.loc[tiered_mask, sat_nodes]
  tiers_sat_nodes_mask['satTrafficList']=[[]]*len(tiers_sat_nodes_mask)
  tiers_sat_nodes_mask.loc[tiered_mask, 'satTrafficList'] = sat_nodes_only.apply(lambda x: sorted(sat_nodes_only.columns[x]), axis=1)
  tiers_sat_nodes_mask['Sat_Node']=['']*len(tiers_sat_nodes_mask)
  tiers_sat_nodes_mask.loc[tiered_mask, 'Sat_Node'] = tiers_sat_nodes_mask[tiered_mask].apply(lambda x: ", ".join(x['satTrafficList']), axis=1)
  tiers_sat_nodes_mask[NonMetricName.SI_SAT_TIER] = [-1]*len(tiers_sat_nodes_mask)
  tiers_sat_nodes_mask.loc[tiered_mask, NonMetricName.SI_SAT_TIER] = tiers_sat_nodes_mask.loc[tiered_mask,'Tier'].astype('int32')
  tiers_sat_nodes_mask['Tier']=['']*len(tiers_sat_nodes_mask)
  tiers_sat_nodes_mask.loc[tiered_mask, 'Tier'] = tiers_sat_nodes_mask.loc[tiered_mask, NonMetricName.SI_SAT_TIER].astype(str)
  tiers_sat_nodes_mask[NonMetricName.SI_CLUSTER_NAME]=''
  tiers_sat_nodes_mask.loc[tiered_mask, NonMetricName.SI_CLUSTER_NAME]=tiers_sat_nodes_mask[tiered_mask].apply(lambda x: f"{x['Tier']} {x['Sat_Node']}", axis=1)
  tiers_sat_nodes_mask[NonMetricName.SI_SAT_NODES]=[set(settings.BASIC_NODE_LIST)]*len(tiers_sat_nodes_mask) 
  tiers_sat_nodes_mask.loc[tiered_mask, NonMetricName.SI_SAT_NODES]=satTrafficListToSiSatNodes(settings, tiers_sat_nodes_mask[tiered_mask])

# For each codelets in current_codelets_runs_df, find their cluster
#   Store the name of the cluster to the SI_CLUSTER_NAME column
#   Also return the a data frame containing by appending all dataframe of the clusters annotated with their names
def do_sat_analysis(optimal_data_df, testSetDF, settings=SatAnalysisSettings()):
    compute_capacities(settings, optimal_data_df)

    # If SI Analysis is disabled
    # Creating an empty Dataframe with column names only SI_SAT_NODES and SI_CLUSTER_NAME
    if settings.disable:
      testSetDF[NonMetricName.SI_CLUSTER_NAME] = ''
      testSetDF[NonMetricName.SI_SAT_NODES] = [chosen_node_set]*len(testSetDF)
      return pd.DataFrame(columns=testSetDF.columns), testSetDF, None

    compute_sw_bias(optimal_data_df)
    compute_sw_bias(testSetDF)

    # Do tiering of training set returning tiering table and cluster information.  Cluster info will also be added in place to optimal_data_df
    tiering_table, cluster_info = tier_training_set(settings, optimal_data_df)

    all_test_codelets = lookup_testing_set(settings, testSetDF, tiering_table, cluster_info)

    # compute SI related metric, so return training codelet being matched to testing set in all_clusters
    all_test_codelets, all_clusters = compute_si_based_metrics(settings, all_test_codelets, optimal_data_df)

    # Compute the statistics results and return the resulting summary dataframe
    # This step is optional and can be skipped by setting settings.compute_stats to False
    result_df = compute_stats(settings, all_test_codelets)

    return all_clusters, all_test_codelets, result_df, tiering_table

def get_chosen_node_set(settings):
  return (set(settings.BASIC_NODE_LIST) |  {settings.CU_NODE_DICT[n] for n in settings.CU_NODE_SET}) & settings.chosen_node_set
  
def compute_si(settings, all_clusters, clustered_test_df):
  norm = "row"
  return compute_only(all_clusters, norm, clustered_test_df, get_chosen_node_set(settings))

def make_si_data(settings, all_clusters, clustered_test_df):
  norm = "row"
  return mk_si_data(all_clusters, norm, clustered_test_df, get_chosen_node_set(settings))
  
def get_relevant_codelets(optimal_data_df, all_test_codelets):
  peer_mask = optimal_data_df[NonMetricName.SI_CLUSTER_NAME].isin(set(all_test_codelets[NonMetricName.SI_CLUSTER_NAME].to_list())-set({''}))
  all_clusters = optimal_data_df[peer_mask].copy()
  clustered_test_df = all_test_codelets[all_test_codelets[NonMetricName.SI_CLUSTER_NAME] != ''].copy()
  return all_clusters, clustered_test_df
  
def compute_si_based_metrics(settings, all_test_codelets, optimal_data_df):
  all_clusters, clustered_test_df = get_relevant_codelets(optimal_data_df, all_test_codelets)
  # all_clusters, _, clustered_test_df = compute_only(all_clusters, norm, clustered_test_df, 
  #                                                   set(BASIC_NODE_LIST) |  {CU_NODE_DICT[n] for n in CU_NODE_SET & set(settings.ALL_NODE_LIST)})
  all_clusters, _, clustered_test_df = compute_si(settings, all_clusters, clustered_test_df)

  added_columns = set(clustered_test_df.columns)-set(all_test_codelets.columns)
  all_test_codelets = pd.merge(left=all_test_codelets, right=clustered_test_df[KEY_METRICS+sorted(added_columns)], on=KEY_METRICS, how='outer')
  # Compute min and max for S and I for each cluster 
  cluster_groups=all_clusters[[NonMetricName.SI_CLUSTER_NAME, 'Intensity', 'Saturation']].groupby(
    NonMetricName.SI_CLUSTER_NAME).agg({'Intensity': ['min', 'max'], 'Saturation': ['min', 'max']})
  # Flattern the columns
  cluster_groups.columns=[' '.join(col).strip().replace(' ','_') for col in cluster_groups.columns.values]
  cluster_groups['s_length'] = round(cluster_groups['Saturation_max'] - cluster_groups['Saturation_min'],2)
  cluster_groups['i_length'] = round(cluster_groups['Intensity_max'] - cluster_groups['Intensity_min'],2)
  cluster_groups['s_ratio'] = round(cluster_groups['Saturation_max'] / cluster_groups['Saturation_min'],2)
  cluster_groups['i_ratio'] = round(cluster_groups['Intensity_max'] / cluster_groups['Intensity_min'],2)

  #cluster_groups['Norm_Tier'] = codelet_tier + ((peer_codelet_df['Saturation'] - peer_codelet_df['Saturation'].min())/s_length)
  cluster_groups['Box_Length'] = '{' + cluster_groups['s_length'].astype(str) + ' , ' + cluster_groups['i_length'].astype(str) + '}'
  cluster_groups['Box_Ratio'] = '{' + cluster_groups['s_ratio'].astype(str) + ' , ' + cluster_groups['i_ratio'].astype(str) + '}'
  # Flattern the groupped df
  #cluster_groups = cluster_groups.stack()[['Box_Length', 'Box_Ratio']]
  all_test_codelets=pd.merge(left=all_test_codelets, right=cluster_groups, on=[NonMetricName.SI_CLUSTER_NAME], how='outer')

  all_test_codelets[NonMetricName.SI_TIER_NORMALIZED] = all_test_codelets[NonMetricName.SI_SAT_TIER] + \
    (all_test_codelets['Saturation']-all_test_codelets['Saturation_min']).clip(lower=0)/all_test_codelets['s_length']

  # Also compute normalized tier for training codelets.  Note that this computation could have done after tiering 
  # but that requires another SI computation so do this here instead.
  all_clusters=pd.merge(left=all_clusters, right=cluster_groups, on=[NonMetricName.SI_CLUSTER_NAME], how='outer')
  all_clusters[NonMetricName.SI_TIER_NORMALIZED] = all_clusters[NonMetricName.SI_SAT_TIER] + \
    (all_clusters['Saturation']-all_clusters['Saturation_min'])/all_clusters['s_length']
  return all_test_codelets, all_clusters

def compute_stats(settings, all_test_codelets):
  if not settings.compute_stats:
    return None

  #all['Variant'] = ''
  result_df = all_test_codelets[['peer_codelet_cnt', 'Tier', 'Sat_Node', 'Sat_Range' ]].copy()

  result_df[MetricName.SHORT_NAME] = all_test_codelets[MetricName.SHORT_NAME] if MetricName.SHORT_NAME in all_test_codelets.columns else all_test_codelets[MetricName.NAME]
  result_df['Variant'] = all_test_codelets['Variant'] if 'Variant' in all_test_codelets.columns else ''

  noSatMask = all_test_codelets['Saturation'].isna()
  all_test_codelets.loc[noSatMask,'SI_Result'] = 'No Cluster'
  no_cluster = len(all_test_codelets[noSatMask])
  codelet_tested = len(all_test_codelets)

  # Compare a mask to pass SI test if saturation within min/max bound of S and I.
  # between() is inclusive <= .. <= by default
  inBoxMask = all_test_codelets['Saturation'].between(all_test_codelets['Saturation_min'], all_test_codelets['Saturation_max']) & \
    all_test_codelets['Intensity'].between(all_test_codelets['Intensity_min'], all_test_codelets['Intensity_max'])
  all_test_codelets.loc[~noSatMask & inBoxMask, 'SI_Result']='Inside Box'
  all_test_codelets.loc[~noSatMask & ~inBoxMask, 'SI_Result']='Outside Box'
  pass_counts = all_test_codelets['SI_Result'].value_counts().reindex(['Inside Box', 'Outside Box'], fill_value=0)
  si_passed = pass_counts['Inside Box']
  si_failed = pass_counts['Outside Box']

  result_df['GFlops']=round(all_test_codelets[MetricName.RATE_FP_GFLOP_P_S], 2)
  result_df['Saturation'] = round(all_test_codelets['Saturation'], 2)
  result_df['Intensity'] = round(all_test_codelets['Intensity'], 2)
  result_df[NonMetricName.SI_CLUSTER_NAME] = all_test_codelets[NonMetricName.SI_CLUSTER_NAME]

  result_df['neg_bias[CNVT,DIV,clu_score,Rec,RHS]']=all_test_codelets[['Nd_CNVT_OPS', 'Nd_DIV_OPS', 'Nd_clu_score', 
                                                                       'Nd_Recurrence', 'Nd_RHS']].values.tolist()
  result_df['pos_bias[VEC_OPS,ISA_EXT,FMA_OPS]'] =all_test_codelets[['Nd_VEC_OPS', 'Nd_ISA_EXT_TYPE', 'Nd_FMA_OPS']].values.tolist()
  result_df.to_csv('Result_report.csv', index = True, header=True)
  num_tiers = len(result_df['Tier'].unique())
  num_clusters = len(result_df[NonMetricName.SI_CLUSTER_NAME].unique())
  print ("Total No. of codelets tested : ", codelet_tested)
  print ("Total No. of codelets outside SI box: ", si_passed)
  print ("Total No. of codelets inside SI box: ", si_failed)
  print ("Total No. of codelets without Cluster : ", no_cluster)
  print ("Total No. of Tiers : ", num_tiers)
  print ("Total No. of unique clusters : ", num_clusters)
  stat_data = [{'Mem Threshold (final)' : settings.satThreshold, 'CU Threshold' : settings.cuSatThreshold,
          'Codelets Tested' : codelet_tested, 'Passed' : si_passed, 'Failed' : si_failed,
          'No Cluster' : no_cluster, 'Tier Count' : num_tiers, 'Cluster Count' : num_clusters}]
  stats_df = pd.DataFrame(stat_data)
  my_file = Path("./statistcs.csv")
  if my_file.is_file():
     # file exists
     stats_df.to_csv('statistcs.csv', mode='a', header=False, index=False)
  else :
     stats_df.to_csv('statistcs.csv', mode='w', header=True, index=False)
  return result_df

def lookup_testing_set(settings, testSetDF, tiering_table, cluster_info):
  tiering_metrics = settings.tiering_metrics
  mem_traffic = settings.mem_traffic
  non_traffic_tiering_metrics = settings.non_traffic_tiering_metrics

  # Do a cross join to form a big table to associate each test codelet to each row of the tiering table.
  # This allows us to compare the tiering metrics of all test codelets with all tiers in the tiering table in parallel (i.e. table lookup)
  test_and_tiering = crossjoin(testSetDF, tiering_table, suffixes=("", "_threshold"))
  # Compare against thresholds.  Need to copy out and rename to matching columns to be able to do the big compare operation.
  tiering_threshold_names = [n+"_threshold" for n in tiering_metrics]
  tiering_thresholds = test_and_tiering[tiering_threshold_names]
  # Renamed the '*_threshold' to original node name so we can do the ">" comparison between two dataframe.
  # Pandas will match the node by column name automatically.
  tiering_thresholds = tiering_thresholds.rename(columns=dict(zip(tiering_threshold_names, tiering_metrics)))
  tiering_sat_checks = test_and_tiering[tiering_metrics] > tiering_thresholds

  # For traffic compute mask to ignore small traffic and also threshold based check
  small_traffic_thresholds = 0.1 * test_and_tiering[mem_traffic].max(axis=1)
  small_traffic_checks = test_and_tiering[mem_traffic].gt(small_traffic_thresholds, axis=0)
  traffic_sat_checks = tiering_sat_checks[mem_traffic] & small_traffic_checks
  traffic_sat_checks_any = traffic_sat_checks.any(axis=1)

  # For non traffic, just use theshold based check
  non_traffic_sat_checks = tiering_sat_checks[non_traffic_tiering_metrics]
  non_traffic_sat_checks_any = non_traffic_sat_checks.any(axis=1)
  test_and_tiering['sat_check'] = traffic_sat_checks_any | non_traffic_sat_checks_any
  # Following using a groupby() call to find, for each codelet, the minimum tier number with sat_check being True, set to np.nan if not found.
  tiers = test_and_tiering[KEY_METRICS+['Tier', 'sat_check']].groupby(KEY_METRICS).apply(lambda x: min(x.loc[x['sat_check'],'Tier'], default=np.nan))

  # Join back to get the rows with right tiers only
  test_and_sats = pd.concat([test_and_tiering[KEY_METRICS+['Tier']],traffic_sat_checks, non_traffic_sat_checks], axis=1)
  test_and_sats = pd.merge(left=test_and_sats, right=tiers.to_frame('Tier'), on=KEY_METRICS+['Tier'], how='right')
  compute_cluster_names(settings, test_and_sats, list(traffic_sat_checks.columns)+list(non_traffic_sat_checks.columns))

  # With cluster name, we can merge with cluster_info to get training info associated
  # Note that both test df and cluster info has 'Sat_Node', 'Tier' and 'SI_SAT_TIER' columns, below we join with both cluster name and sat node expecting them to be consistent 
  test_and_sats = pd.merge(left=test_and_sats, right=cluster_info, on=[NonMetricName.SI_CLUSTER_NAME, 'Sat_Node', 'Tier', NonMetricName.SI_SAT_TIER], how='left')
  tiered_mask = ~test_and_sats['Tier'].isna()
  test_and_sats.loc[tiered_mask, NonMetricName.SI_SAT_TIER]=test_and_sats.loc[tiered_mask, NonMetricName.SI_SAT_TIER].astype('int32')
  test_and_sats.loc[tiered_mask,'Tier']=test_and_sats.loc[tiered_mask,'Tier'].astype(str)
  # Set Cluster name to '' for tiny clusters
  test_and_sats.loc[(test_and_sats['peer_codelet_cnt']<2) | (test_and_sats['peer_codelet_cnt'].isna()), NonMetricName.SI_CLUSTER_NAME] = ''
  # Merge back testSetDF for added columns
  return pd.merge(left=testSetDF, right=test_and_sats[KEY_METRICS+sorted(set(test_and_sats.columns) -set(testSetDF.columns))], on=KEY_METRICS)

def tier_training_set(settings, satSetDF):
  tiering_metrics = settings.tiering_metrics
  mem_fp_reg_rate_metrics = settings.mem_traffic+settings.fp_reg_rate_metrics

  tiering_table = pd.DataFrame()
  sat_table = satSetDF[KEY_METRICS].copy()
  # naMask is True for items not tiered yet so need to be considered
  satSetDF['Tier'] = None 
  gap = None
  tier = 0
  candidate_mask = satSetDF['Tier'].isna()
  uniform_gap = compute_tier_gap(settings, satSetDF, mem_fp_reg_rate_metrics, settings.percentsToCheck) if settings.uniform_gap else None
  while candidate_mask.any():
    curSatSetDF = satSetDF[candidate_mask]
    tier = tier + 1
    settings.satThreshold = 0.1 * tier
    thresholds = compute_node_thresholds(settings, curSatSetDF, mem_fp_reg_rate_metrics, 
                                         settings.percentsToCheck, uniform_gap)
    thresholds['Tier'] = tier
    tiering_table = tiering_table.append(thresholds, ignore_index=True)

    # Compute a mask to record which nodes are saturated 
    sat_mask = (curSatSetDF[tiering_metrics] > thresholds[tiering_metrics])
    # passMask records the rows in the candidate rows passing the saturation test (existence of any saturated node)
    passMask = sat_mask.any(axis=1)
    # update satSetDF in place.  To do this we get the mask to do the update
    # First get a copy of current na mask reflecting this iteration's candidate rows.  
    # The size of this mask is always the same as satSetDF.
    mask_for_satSetDF = candidate_mask.copy()
    # Among the rows with naMask==True, set the True/False value following the pass_mask, 
    # Esseential this operation bring the saturation result back to the overall satSetDF mask,
    # so satSetDF[mask] are the rows passing saturation test.
    mask_for_satSetDF[candidate_mask]=passMask

    satSetDF.loc[mask_for_satSetDF, 'Tier'] = tier
    sat_table.loc[mask_for_satSetDF, 'Tier'] = tier
    sat_table.loc[mask_for_satSetDF, sat_mask.columns] = sat_mask

    candidate_mask = satSetDF['Tier'].isna()
  satSetDF[NonMetricName.SI_SAT_TIER] = satSetDF['Tier'] .astype(int)
  tiering_table[NonMetricName.SI_SAT_TIER] = tiering_table['Tier'] .astype(int)
  all_nodes = set(sat_table)-set(KEY_METRICS+['Tier'])
  compute_cluster_names(settings, sat_table, all_nodes)
  # Since sat_table and satSetDF are in the same index order, so just assign the cluster name column directly
  columns_to_add = sorted(set(sat_table.columns)-set(satSetDF.columns))
  satSetDF[columns_to_add] = sat_table[columns_to_add]

  # Compute cluster information using tierred results.
  cluster_info = compute_cluster_info(satSetDF, all_nodes)

  # Counting only for tiering
  tiering_table['Training Set Count']=tiering_table.Tier.map(satSetDF.Tier.value_counts())
  return tiering_table, cluster_info 

# Givebn tiered training set result, compute aggregated information for the cluster
def compute_cluster_info(satSetDF, all_nodes):
  # Now compute the Sat_Range.  Including 'Sat_Node' instead of satTrafficList because list cannob be used in groupby
  # Also reset_index() to bring the cluster name and Sat_Node to column names.  Including Tier and SI_SAT_TIER assuming they are consistent with cluster name which is the primary key
  cluster_info = satSetDF.groupby([NonMetricName.SI_CLUSTER_NAME, 'Sat_Node', 'Tier', NonMetricName.SI_SAT_TIER]).agg({ n: ['min', 'max'] for n in all_nodes}).reset_index()
  # TODO : fix types more cleanly
  cluster_info[NonMetricName.SI_SAT_TIER]=cluster_info[NonMetricName.SI_SAT_TIER].astype(int)
  cluster_info['Tier']=cluster_info['Tier'].astype(str)
  # Flattern the columns
  cluster_info.columns=[' '.join(col).strip().replace(' ','_') for col in cluster_info.columns.values]
  cluster_info['Sat_Range'] = cluster_info.apply(lambda x: ', '.join([f"{elem}: [{x[elem+'_max']}  {x[elem+'_min']}]" for elem in x['Sat_Node'].split(', ')]), axis=1)

  # Simple counting of codelets in each cluster
  cluster_counts = satSetDF[NonMetricName.SI_CLUSTER_NAME].value_counts()
  # Convert the counts to dataframe to join with cluster_info and return
  cluster_counts = cluster_counts.to_frame('peer_codelet_cnt').reset_index().rename(columns={'index': NonMetricName.SI_CLUSTER_NAME})
  return pd.merge(left=cluster_info, right=cluster_counts, on=NonMetricName.SI_CLUSTER_NAME, how='left')

    
# Some leftover routines not removed but unused.
################################################################################
# setting up excelsheet to color
################################################################################
def load_workbook(tier_book_path):
   if os.path.exists(tier_book_path):
      return openpyxl.load_workbook(tier_book_path)
   return  openpyxl.Workbook()


################################################################################
# Coloring
################################################################################

opRed = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
opYellow = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
opGreen = openpyxl.styles.PatternFill(start_color="008000", end_color="008000", fill_type="solid")
opBlue = openpyxl.styles.PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
opPurple = openpyxl.styles.PatternFill(start_color="800080", end_color="800080", fill_type="solid")
# Find max in traffic columns + perfcent columns, save to maxDict
def createSubcluster(settings, data, subSatList, short_name):
  # get rows that we need to check
  rowsToCheck = data.index.tolist()
  initDict(maxOfColumn, subNodeTrafficToCheck)
  target_df = pd.DataFrame(columns=data.columns.tolist())
  csv_string = short_name + '_sub_cluster.csv'
  #print (data)
  if not subSatList:
      for row in rowsToCheck:
        #print (row)
        #print ("No sub sat nodes")
        for column in subNodeTrafficToCheck:
          columnIndex = data.columns.get_loc(column)
          # init list of blocks to color 
          num = data.iloc[row, columnIndex]
          if num <= 50:
              maxOfColumn[column].append(row)
  else:
      for row in rowsToCheck:
        #print (row)
        #print (subSatList)
        for column in subNodeTrafficToCheck and subSatList:
          # init list of blocks to color 
          columnIndex = data.columns.get_loc(column)
          num = data.iloc[row, columnIndex]
          if num >= 50:
              maxOfColumn[column].append(row)
    #end columnfor
  #end rowfor
  codelet_list = maxOfColumn[column]
  if not subSatList:
      for column in settings.percentsToCheck:
          codelet_list = list(set(codelet_list) & set(maxOfColumn[column]))
      for row in codelet_list:
          target_df = target_df.append(data.iloc[row], ignore_index=True)
  else :
      codelet_list = maxOfColumn[subSatList[0]]
      for column in settings.percentsToCheck and subSatList:
          codelet_list = list(set(codelet_list) & set(maxOfColumn[column]))
      for row in codelet_list:
          target_df = target_df.append(data.iloc[row], ignore_index=True)
  return target_df


def concat_ordered_columns(frames):
    columns_ordered = []
    for frame in frames:
        columns_ordered.extend(x for x in frame.columns if x not in columns_ordered)
    final_df = pd.concat(frames, ignore_index=True)
    return final_df[columns_ordered]

def find_swbias_cluster(satSetDF, testDF, swbias_threshold):
    mask = satSetDF[NonMetricName.SI_SW_BIAS] >= swbias_threshold
    if testDF.iloc[0]['Net_SW_Bias'] >= swbias_threshold:
         return satSetDF[mask]
    else:
         return satSetDF[~mask]

def do_swbias_clustering(settings, peer_codelet_df, testDF, satTrafficList):
    #sw_bias_df = compute_sw_bias(peer_codelet_df)
    #peer_codelet_swBias_df = pd.merge(peer_codelet_df, sw_bias_df)

    # sw_bias_df = compute_sw_bias(testDF)
     # get SW bias Vector
    neg_bias_vec =[testDF['Nd_CNVT_OPS'].item(), testDF['Nd_DIV_OPS'].item(), testDF['Nd_clu_score'].item(),
                   testDF['Nd_Recurrence'].item(), testDF['Nd_RHS'].item()] 
    pos_bias_vec =[testDF['Nd_VEC_OPS'].item(), testDF['Nd_ISA_EXT_TYPE'].item(), testDF['Nd_FMA_OPS'].item()]

    # test_swBias_df = pd.merge(testDF, sw_bias_df)
    swbias_cluster_df = find_swbias_cluster(peer_codelet_df, testDF, 0)
    cdlt_count = swbias_cluster_df.shape[0]
    chosen_node_set = set(settings.BASIC_NODE_LIST)
    short_name = str(testDF.iloc[0][MetricName.SHORT_NAME])
    if cdlt_count >= 3:
        for elem in satTrafficList:
            if elem in settings.CU_NODE_SET:
                #print (CU_NODE_DICT[elem])
                chosen_node_set.add(CU_NODE_DICT[elem])
        outputfile = short_name + '_sw_bias_'+ "_SI.csv"
        norm = "row"
        title = "SI"
        target_df = pd.DataFrame()
        #compute_and_plot('XFORM', swbias_cluster_df, outputfile, norm, title, chosen_node_set, target_df)
        my_cluster_df, my_cluster_and_test_df, my_test_df = compute_only(swbias_cluster_df, norm, testDF, chosen_node_set)
        peer_dfs = [swbias_cluster_df,testDF]
        final_df = concat_ordered_columns(peer_dfs)
        final_df.to_csv(outputfile, index = False, header=True)
        #bias_result = test_and_plot_orig('ORIG', final_df, outputfile, norm, title, chosen_node_set, target_df, short_name)
        bias_result = True
        s_length = swbias_cluster_df['Saturation'].max() - swbias_cluster_df['Saturation'].min()
        i_length = swbias_cluster_df['Intensity'].max() - swbias_cluster_df['Intensity'].min()
        box_length = '{' + str(round(s_length, 2)) + ' , ' + str(round(i_length, 2)) + '}'

        s_ratio = swbias_cluster_df['Saturation'].max() / swbias_cluster_df['Saturation'].min()
        i_ratio = swbias_cluster_df['Intensity'].max() / swbias_cluster_df['Intensity'].min()
        box_ratio = '{' + str(round(s_length, 2)) + ' , ' + str(round(i_length, 2)) + '}'

        if bias_result == True :
            print (short_name, " swbias cluster Outside the SI Box =>")
            result_df = pd.DataFrame({MetricName.SHORT_NAME : short_name, 'peer_codelet_cnt' : cdlt_count,'SI_Result' : 'Outside Box',
                        'Box_Length' : box_length, 'Box_Ratio' : box_ratio, 'Neg_SW_bias_Vec' : str(neg_bias_vec), 'Pos_SW_bias_Vec' : str(pos_bias_vec)}, index=[0])
        else:
            print (short_name, " swbias cluster Inside the SI Box =>")
            result_df = pd.DataFrame({MetricName.SHORT_NAME : short_name, 'peer_codelet_cnt' : cdlt_count,'SI_Result' : 'Inside Box',
                        'Box_Length' : box_length, 'Box_Ratio' : box_ratio, 'Neg_SW_bias_Vec' : str(neg_bias_vec), 'Pos_SW_bias_Vec' : str(pos_bias_vec)}, index=[0])
    else:
        print (short_name, " swbias cluster not present =>")
        result_df = pd.DataFrame({MetricName.SHORT_NAME : short_name, 'peer_codelet_cnt' : 0,'SI_Result' : 'No Cluster',
        'Box_Length' : 0, 'Box_Ratio' : 0, 'Neg_SW_bias_Vec' : str(neg_bias_vec), 'Pos_SW_bias_Vec' : str(pos_bias_vec)}, index=[0])
    return result_df



def print_coloured_tiers(settings, short_name, codelet_tier, full_df):
    tier_book_path = short_name + "_tier.xlsx"
    tier_book = load_workbook(tier_book_path);
    highlightSheet = tier_book.create_sheet(short_name + "_" + str(codelet_tier), 0)
    sheet_title = short_name + "_" + str(codelet_tier) + "_" + "Highlights"
        # import all rows into xlsx package
    for r in dataframe_to_rows(full_df, index=False, header=True):
        highlightSheet.append(r)
    # Commented out following lines as it seems to be possible using Pandas library to compute max somehow.
    #findMaxInColumnsToColor(full_df, settings.non_cu_tiering_metrics, settings.percentsToCheck)
    #colorMaxInColumn(full_df, coloured_maxOfColumn, opBlue, highlightSheet)
    tier_book.save(tier_book_path)


def compute_capacities(settings, df):
    #chosen_node_set = ALL_NODE_SET
    chosen_node_set = settings.chosen_node_set
    nodes_without_units = {n.split(" ")[0] for n in chosen_node_set} 
    CapacityData(df).set_chosen_node_set(nodes_without_units).compute()

def tiering_only(args):
    NUM_POINTS=3
    SMALL = np.finfo(float).eps * 100
    SMALL = 0.001
    training_set_csv = args.training_csv_file
    chosen_node_set = ALL_NODE_SET
    settings = SatAnalysisSettings(satThreshold=float(args.sat_threshold), cuSatThreshold=float(args.cu_sat_threshold), chosen_node_set=chosen_node_set) 
    optimal_data_df = pd.read_csv(training_set_csv)

    compute_capacities(settings, optimal_data_df)

    tiering_table, cluster_info = tier_training_set(settings, optimal_data_df)
    min_peer_mask = (cluster_info['peer_codelet_cnt']>=2)
    cluster_info = cluster_info[min_peer_mask]
    cluster_with_tier_thresholds = pd.merge(left=cluster_info, 
                                            right=tiering_table.rename(columns=dict(zip(settings.tiering_metrics, [m+'_threshold' for m in settings.tiering_metrics]))), 
                                            on=[NonMetricName.SI_SAT_TIER])
    cluster_with_tier_thresholds['satTrafficList'] = cluster_with_tier_thresholds['Sat_Node'].str.split(', ')
    cluster_with_tier_thresholds[NonMetricName.SI_SAT_NODES]=satTrafficListToSiSatNodes(settings, cluster_with_tier_thresholds)
    cluster_with_tier_thresholds['nonSatTrafficList'] = cluster_with_tier_thresholds.apply(lambda x: [n for n in settings.tiering_metrics if n not in x['satTrafficList']], axis=1)
    cluster_with_tier_thresholds['sat_lb'] = cluster_with_tier_thresholds.apply(lambda x: {n:x[n+'_threshold'] for n in x['satTrafficList']}, axis=1)
    cluster_with_tier_thresholds['sat_ub'] = cluster_with_tier_thresholds.apply(lambda x: {n:x[n+'_max'] for n in x['satTrafficList']}, axis=1)
    cluster_with_tier_thresholds['nonsat_lb'] = cluster_with_tier_thresholds.apply(lambda x: {n:0 for n in x['nonSatTrafficList']}, axis=1)
    cluster_with_tier_thresholds['nonsat_ub'] = cluster_with_tier_thresholds.apply(lambda x: {n:x[n+'_threshold'] for n in x['nonSatTrafficList']}, axis=1)
    cluster_with_tier_thresholds['lb'] = cluster_with_tier_thresholds.apply(lambda x: {**x['sat_lb'], **x['nonsat_lb']}, axis=1)
    cluster_with_tier_thresholds['ub'] = cluster_with_tier_thresholds.apply(lambda x: {**x['sat_ub'], **x['nonsat_ub']}, axis=1)
    cluster_with_tier_thresholds['range'] = cluster_with_tier_thresholds.apply(lambda x: {n: sorted(pd.unique(np.linspace(x['lb'][n]+SMALL, x['ub'][n], NUM_POINTS))) for n in settings.tiering_metrics}, axis=1)
    # See: https://stackoverflow.com/questions/38231591/split-explode-a-column-of-dictionaries-into-separate-columns-with-pandas
    # Using json_normalize to expand the dictionary to multiple columns
    cluster_with_tier_thresholds = pd.concat([cluster_with_tier_thresholds, pd.json_normalize(cluster_with_tier_thresholds['range'])], axis=1)
    # For each metric expanding from the list to real rows so metric of [1,2,3] will be expanded into 3 rows with 1, 2 and 3 as metric value while keeping the rest the same
    # See: https://stackoverflow.com/questions/12680754/split-explode-pandas-dataframe-string-entry-to-separate-rows
    for m in settings.tiering_metrics:
      cluster_with_tier_thresholds = cluster_with_tier_thresholds.explode(m)
      
    plot_data = cluster_with_tier_thresholds.reset_index(drop=True)
    # Currently these metrics seems to be unused but set to zero for capacity calculation
    plot_data[MetricName.NAME] = 'plot-data'
    plot_data[MetricName.TIMESTAMP] = list(range(0, len(plot_data)))

    plot_data1=plot_data[KEY_METRICS+settings.tiering_metrics].copy()
    plot_data1[[MetricName.STALL_RS_PCT, MetricName.STALL_FE_PCT]] = 0
    compute_capacities(settings, plot_data1)

    plot_data1 = lookup_testing_set(settings, plot_data1, tiering_table, cluster_info)
    plot_data1 = plot_data1[plot_data1[NonMetricName.SI_CLUSTER_NAME] != '']
    #foo=pd.merge(left=plot_data[KEY_METRICS+['SiClusterName']], right=plot_data1[KEY_METRICS+['SiClusterName']], on=KEY_METRICS)
    #no_mask=(foo['SiClusterName_x'] != foo['SiClusterName_y'])
    #no_mask = (foo['SiClusterName_y'] == '')
    #bar=pd.DataFrame([plot_data1.loc[no_mask,KEY_METRICS+['SiClusterName']+settings.tiering_metrics].iloc[0,:]])
    #all_bar = lookup_testing_set(settings, bar, tiering_table, cluster_info)

    # chosen_node_set = ALL_NODE_SET
    # nodes_without_units = {n.split(" ")[0] for n in chosen_node_set} 
    # CapacityData(plot_data).set_chosen_node_set(nodes_without_units).compute()
    all_clusters, plot_data1 = get_relevant_codelets(optimal_data_df, plot_data1)
    all_clusters, _, clustered_test_df = compute_si(settings, all_clusters, plot_data1)
    all_clusters['is_cluster'] = True
    clustered_test_df['is_cluster'] = False

    combined = pd.concat([all_clusters, clustered_test_df], axis=0)
    combined[MetricName.SHORT_NAME] = ''
    # Free up some memory before plotting
    del cluster_with_tier_thresholds
    del plot_data
    del plot_data1
    del all_clusters
    del clustered_test_df
    gc.collect()
    print("Generate SI plots")

    combined.groupby(NonMetricName.SI_CLUSTER_NAME).apply(lambda x: 
      plot_si(settings, x[x['is_cluster']], x[~x['is_cluster']], training_set_csv, 
              os.path.join('c:/temp', f'{x.name}.png'.replace(', ', ';').replace('/', '_').replace(' ', '-'))))
    pass

def plot_si(settings, cluster_df, test_df, training_csv_filename, outfile):
    print(f'Plotting graph to: {outfile}')
    siData = make_si_data(settings, cluster_df, test_df)
    siPlot = StaticSiPlot(siData, training_csv_filename, outfile)
    siPlot.skip_adjustText()
    siPlot.hide_labels()
    #siPlot.set_labels([MetricName.RATE_FP_GFLOP_P_S, 'SI'])
    siPlot.compute_and_plot()


def full_analysis(args):
    # csv to read should be first argument
    csvToRead = args.training_csv_file
    csvTestSet = args.test_csv_file
    #inputfile = []
    #sys.setrecursionlimit(10**9) 
  # if 3 arg specified, assumes 3rd is threshold replacement
    #print("No of sys args : ", len(sys.argv))
    chosen_node_set = ALL_NODE_SET
    settings = SatAnalysisSettings(satThreshold=float(args.sat_threshold), cuSatThreshold=float(args.cu_sat_threshold), chosen_node_set=chosen_node_set) 

    print("Attempting to read", csvToRead)

    # read into pandas
    mainDataFrame = pd.read_csv(csvToRead)
    TestSetDF = pd.read_csv(csvTestSet)
    grouped = TestSetDF.groupby(MetricName.VARIANT)
    #mask = (TestSetDF['Set'] == 'BENEFITTING') & (TestSetDF['Variant'] == 'ORIG')
    #mask = (TestSetDF['Set'] == 'BENEFITTING')

    print("Read successful!")

    # save original dataframe
    originalDataFrame = mainDataFrame

    # SIMD_RATE divided by max traffic column for memory
    addAfterColumn = mainDataFrame.columns.get_loc(MetricName.RATE_RAM_GB_P_S) + 1
    mainDataFrame.insert(addAfterColumn, "SIMD_MEM_Intensity",
      mainDataFrame[MetricName.RATE_REG_SIMD_GB_P_S] / mainDataFrame[[MetricName.RATE_L1_GB_P_S, MetricName.RATE_L2_GB_P_S, MetricName.RATE_L3_GB_P_S, MetricName.RATE_RAM_GB_P_S]].max(axis=1))

    # FLOP_RATE divided by max traffic column for memory
    addAfterColumn = mainDataFrame.columns.get_loc(MetricName.RATE_FP_GFLOP_P_S) + 1
    mainDataFrame.insert(addAfterColumn, "FLOP_MEM_Intensity",
    mainDataFrame[MetricName.RATE_FP_GFLOP_P_S] / mainDataFrame[[MetricName.RATE_L1_GB_P_S, MetricName.RATE_L2_GB_P_S, MetricName.RATE_L3_GB_P_S, MetricName.RATE_RAM_GB_P_S]].max(axis=1)/8)
    if RUN_SI:
      #do_sat_analysis(mainDataFrame, TestSetDF[mask])
      nodes_without_units = {n.split(" ")[0] for n in chosen_node_set} 
      CapacityData(TestSetDF).set_chosen_node_set(nodes_without_units).compute()
      all_clusters, all_test_codelets, stats_df, tiering_table = do_sat_analysis(mainDataFrame, TestSetDF, settings)
      tiering_table = tiering_table[settings.tiering_metrics+[NonMetricName.SI_SAT_TIER]]
      # Rename tiering metric columns before joining to avoid conflicts
      tiering_table = tiering_table.rename(columns={f:f'Tiering_{f}' for f in settings.tiering_metrics})
      all_test_codelets = pd.merge(left=all_test_codelets, right=tiering_table, on=[NonMetricName.SI_SAT_TIER], how='inner')
      all_test_codelets.to_csv(args.out_file, index = False, header=True);
    # if PRINT_ALL_CLUSTERS:
    #   find_all_clusters(mainDataFrame)


if __name__ == "__main__":
    parser = ArgumentParser(description='Sat Analysis: compute cluster, S and I for test codelets.')
    parser.add_argument('-m', nargs='?', help='the training (modeling) set csv file', required=True, dest='training_csv_file')
    parser.add_argument('-t', nargs='?', help='the test set csv file(s)', required=False, dest='test_csv_file')
    parser.add_argument('-o', nargs='?', default='SI_out.csv', help='the output csv file (default SI_out.csv)', dest='out_file')
    parser.add_argument('--tiering-only', action='store_true', help='Only perform tiering', dest='tiering_only')
    parser.add_argument('--cu-sat-threshold', nargs='?', help='CU saturation threshold (0-1)', dest='cu_sat_threshold', default=SatAnalysisSettings.CU_SAT_THRESHOLD_DEFAULT)
    parser.add_argument('--sat-threshold', nargs='?', help='non-CU saturation threshold (0-1)', dest='sat_threshold', default=SatAnalysisSettings.SAT_THRESHOLD_DEFAULT)
    args = parser.parse_args()
    # csv to read should be first argument
    #main(sys.argv[1:])
    if args.tiering_only:
      tiering_only(args)
    else:
      full_analysis(args) 
