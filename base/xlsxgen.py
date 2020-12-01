from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

import os, csv
import itertools

class XlsxGenerator:

    def __init__(self):
        self.tiered_metrics = {}
        self.cutoffs = {}
        self.highlights = {}

    def set_header(self, header):
        self.single = header == 'single'

    def set_scheme(self, scheme):
        if (scheme == 'general'):
            import general_scheme
            self.grouping = general_scheme.grouping
            general_scheme.apply(self)

    def put_thresholds(self, metric, thresholds):
        self.tiered_metrics[metric] = thresholds

    def put_cutoff(self, metric, cutoff):
        self.cutoffs[metric] = cutoff

    def put_highlight(self, metric, value):
        self.highlights[metric] = value

    def from_dataframe(self, title, df, outfile):
        _process(self, title, df, outfile)

class DerivedMetric:
    def __init__(self, value, inputs, output):
        self.value = value
        self.inputs = inputs
        self.output = output

    def apply(self, loop):
        inputs = [ loop[x] for x in self.inputs ]
        return self.output(*inputs)

_colors = [ '2EB62C', '83D475', 'C5E8B7' ]
_highlight = 'DAAE5D'

def _try_cast(value):
    try:
        if len(value) == 0:
            return 0
        else:
            return float(value)
    except:
        return value

def _pct_diff(x, y):
    return 2 * abs(x - y) / (x + y)

def _format_value(field, value):
    if field == 'scores':
        if not value or value == '{}':
            return None
        scores = eval(value)
        scores.reverse()
        return '[' + ', '.join('%.3f' % x for x in scores) + ']'
    else:
        return _try_cast(value)

def _onerow_format_field(x):
    if x == 'VecType[Ops]':
        return 'VecType'
    elif x.startswith('%lfb'):
        return x[x.index(':') + 1 :].upper() + '%'
    last = x.rindex('_') if '_' in x else len(x)
    metric = (x[x.index('[') + 1 : x.rindex(']')]) if '[' in x else x[:last]
    metric = 'CluScores' if metric == 'scores' else metric
    units = x[(last + 1) :]
    units = units if not units or units == '%' else (' (%s)' % units)
    return metric + units

def _tworow_format_field(x):
    if x == 'VecType[Ops]':
        return 'VecType'
    else:
        return (
            (x[x.index('[') + 1 : x.rindex(']')])
            if '[' in x
            else (x[: x.rindex('_')] if '_' in x else x)
        )

def _find_tiers(df, field, thresholds):
    lastmax, tiers = max(df[field].values), []
    for threshold in thresholds:
        members = [ idx for idx, value in df[field].iteritems() if ((_pct_diff(value, lastmax) <= threshold) and value <= lastmax) ]
        tiers.append(members)
        flattened = list(itertools.chain(*tiers))
        lastmax = max(value for idx, value in df[field].iteritems() if (idx not in flattened))
    return tiers

def _get_tier(tiers, loop):
    for idx, tier in enumerate(tiers):
        if loop in tier:
            return idx
    return -1

def _color_for_tier(tier):
    return _colors[tier]

def _fill_from_color(rgb):
    return PatternFill('solid', fgColor=rgb)

def _column_for_metric(header, metric):
    return get_column_letter(header.index(metric.value) + 1)

def _color_cell(ws, col, row, color):
    cell = ws[col + str(row)]
    cell.fill = _fill_from_color(color)

def _process(xlsxgen, title, df, outfile):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    curr_row = 0
    def append(row):
        nonlocal curr_row
        ws.append(row)
        curr_row = curr_row + 1
    group_header, data_header, units_header = [], [], []
    grouping = xlsxgen.grouping.copy()
    for group in xlsxgen.grouping:
        # drop empty groups
        if all((x not in df.columns) for x in group[1]):
            grouping.remove(group)
            continue
        group_header.append(group[0])
        for _ in range(len(group[1]) - 1):
            group_header.append('')
        data_header.extend(group[1])
        for metric in filter(lambda x: isinstance(x, DerivedMetric), group[1]):
            df[metric.value] = metric.apply(df).values
    hidden_cols = [
        get_column_letter(i + 1)
        for i, x in enumerate(data_header)
        if isinstance(x, tuple) and x[1]
    ]
    for hidden in hidden_cols:
        ws.column_dimensions[hidden].hidden = True
    data_header = [
        (x[0].value if isinstance(x, tuple) else x if isinstance(x, str) else x.value)
        for x in data_header
    ]
    units_header = [x[x.rindex('_') + 1 :] if '_' in x else None for x in data_header]
    append(group_header)
    start_column = 1
    for group in grouping:
        num_cols = len(group[1])
        ws.merge_cells(
            start_row=1,
            start_column=start_column,
            end_row=1,
            end_column=start_column + num_cols - 1,
        )
        start_column = start_column + num_cols
    append((_onerow_format_field if xlsxgen.single else _tworow_format_field)(x) for x in data_header)
    if not xlsxgen.single:
        append(units_header)
    tiered = {}
    for metric in xlsxgen.tiered_metrics:
        col = _column_for_metric(data_header, metric)
        tiered[col] = _find_tiers(df, metric, xlsxgen.tiered_metrics[metric])
    for idx, loop in df.iterrows():
        append([_format_value(x, loop.get(x, None)) for x in data_header])
        for metric, value in xlsxgen.highlights.items():
            if loop[metric] == value:
                _color_cell(ws, _column_for_metric(data_header, metric), curr_row, _highlight)
        for metric, cutoff in xlsxgen.cutoffs.items():
            if _try_cast(loop[metric]) >= cutoff:
                _color_cell(ws, _column_for_metric(data_header, metric), curr_row, _color_for_tier(0))
        for metric, tiers in tiered.items():
            found = _get_tier(tiers, idx)
            if found >= 0:
                _color_cell(ws, metric, curr_row, _color_for_tier(found))
    wb.save(filename=outfile)