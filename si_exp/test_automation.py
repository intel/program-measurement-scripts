from openpyxl.utils.dataframe import dataframe_to_rows
import sys
sys.path.append('.\test_SI')
from test_SI import compute_and_plot
from test_SI import test_and_plot_orig
from metric_names import MetricName
from metric_names import NonMetricName
from sw_bias import compute_sw_bias
import openpyxl
import openpyxl.chart
import pandas as pd
import numpy as np  
import sys
import importlib
import re
from pathlib import Path
from os import path

# csv to read should be first argument
csvToRead = sys.argv[1]
csvTestSet = sys.argv[2]
# percent within max in column for color
# TODO unused at the moment
satThreshold = 0.10
cuSatThreshold = 0.15

si_passed = 0
si_failed = 0
no_cluster = 0
unique_sat_node_clusters = []
unique_tiers = []
DO_SUB_CLUSTERING = False
DO_DEBUG_LOGS = False
PRINT_ALL_CLUSTERS = False
PRINT_COLOURED_TIERS = False
RUN_SI = True
RUN_SW_BIAS = False

CU_NODE_SET={MetricName.STALL_FE_PCT, MetricName.STALL_LB_PCT, MetricName.STALL_SB_PCT, MetricName.STALL_LM_PCT, MetricName.STALL_RS_PCT}
CU_NODE_DICT={MetricName.STALL_FE_PCT:'FE', MetricName.STALL_LB_PCT:'LB', MetricName.STALL_SB_PCT:'SB', MetricName.STALL_LM_PCT:'LM', MetricName.STALL_RS_PCT:'RS'}

# No Frontend
#CU_NODE_SET={'%lb', '%sb', '%lm', '%rs'}
#CU_NODE_DICT={'%lb':'LB', '%sb':'SB', '%lm':'LM', '%rs':'RS'}

BASIC_NODE_LIST=['L1', 'L2', 'L3', 'FLOP', 'VR', 'RAM']

# memory traffic
trafficToCheck = [ MetricName.RATE_REG_SIMD_GB_P_S, MetricName.RATE_FP_GFLOP_P_S, MetricName.RATE_L1_GB_P_S, MetricName.RATE_L2_GB_P_S, MetricName.RATE_L3_GB_P_S, MetricName.RATE_RAM_GB_P_S ]
memTrafficToCheck = [ MetricName.RATE_REG_SIMD_GB_P_S, MetricName.RATE_L1_GB_P_S, MetricName.RATE_L2_GB_P_S, MetricName.RATE_L3_GB_P_S, MetricName.RATE_RAM_GB_P_S ]
archIntensityToCheck = ["SIMD_MEM_Intensity", "FLOP_MEM_Intensity"]

#ALL_NODE_LIST =  [ "register_simd_rate_gb/s", "flop_rate_gflop/s", "l1_rate_gb/s", "l2_rate_gb/s", "l3_rate_gb/s", "ram_rate_gb/s",
#                   "%sb", "%lm", "%rs", "%lb"]
ALL_NODE_LIST =  [MetricName.RATE_L1_GB_P_S, MetricName.RATE_L2_GB_P_S, MetricName.RATE_L3_GB_P_S, MetricName.RATE_RAM_GB_P_S, MetricName.RATE_REG_SIMD_GB_P_S, MetricName.RATE_FP_GFLOP_P_S,
                            MetricName.STALL_SB_PCT, MetricName.STALL_LM_PCT, MetricName.STALL_LB_PCT]

# arith
#percentsToCheck = ["register_simd_rate_gb/s", "%ops[vec]", "%inst[vec]", "%prf", "%sb", "%rs", "%lb", "%rob", "%lm", "%frontend" ]
#percentsToCheck = [ "%sb", "%lm", "%frontend" , "%rs", "%lb"]
percentsToCheck = [ MetricName.STALL_SB_PCT, MetricName.STALL_LM_PCT, MetricName.STALL_LB_PCT]
cuTrafficToCheck = [ MetricName.STALL_SB_PCT, MetricName.STALL_LM_PCT, MetricName.STALL_LB_PCT]
subNodeTrafficToCheck = [ MetricName.STALL_SB_PCT, MetricName.STALL_LM_PCT, MetricName.STALL_LB_PCT, MetricName.STALL_FE_PCT, MetricName.STALL_RS_PCT]

# this dict contains columns + rows of those columns that need to be colored
# assumption is that you don't add any more rows else dict becomes out of dat
maxOfColumn = {}
coloured_maxOfColumn = {}

# Initializes a list of things in a dict as lists
def initDict(maxOfColumn, listOfColumns):
  for i in listOfColumns:
    maxOfColumn[i] = []

# init lists
initDict(maxOfColumn, trafficToCheck)
initDict(maxOfColumn, percentsToCheck)
initDict(maxOfColumn, archIntensityToCheck)

column_names = [MetricName.SHORT_NAME, 'Variant', 'Set', 'peer_codelet_cnt', 'Tier', 'Sat_Node', 'Sat_Range', 'SI_Result', 'Box_Length', 'Box_Ratio', 'SW_bias_CLS_CDLTS',
               'SW_bias_Result', 'SW_bias_box_length', 'SW_bias_box_ratio', 'Intensity', 'Saturation', 'GFlops',
               'neg_bias[CNVT,DIV,clu_score,Rec,RHS]', 'pos_bias[VEC_OPS,ISA_EXT,FMA_OPS]', 'SW_bias']
result_df = pd.DataFrame(columns=column_names)

################################################################################
# setting up excelsheet to color
################################################################################
def load_workbook(tier_book_path):
   if path.exists(tier_book_path):
      return openpyxl.load_workbook(tier_book_path)
   return  openpyxl.Workbook()


################################################################################
# Coloring
################################################################################

opRed = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
opYellow = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
opGreen = openpyxl.styles.PatternFill(start_color="008000", end_color="008000", fill_type="solid")
opBlue = openpyxl.styles.PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
opPurple = openpyxl.styles.PatternFill(start_color="800080", end_color="800080", fill_type="solid")

# loop through max columns, color them
def colorMaxInColumn(data, columnDict, colorToUse, pyxlSheet):
  for key in columnDict.keys():
    columnIndex = data.columns.get_loc(key)
    rowsToColor = columnDict[key]
    # column + 1 because 1-index
    excelColumn = openpyxl.utils.cell.get_column_letter(columnIndex + 1)
    for row in rowsToColor:
      # row + 2 needed because 0 indexed and first row is column names
      excelRow = row + 2
      cellCoord = excelColumn + str(excelRow)
      activeCell = pyxlSheet[cellCoord]
      # color this cell
      activeCell.fill = opGreen

# Find max in traffic columns + perfcent columns, save to maxDict
def findMaxInColumnsToColor(data, main_traffic, cu_traffic):
  # get rows that we need to check
  rowsToCheck = data.index.tolist()
  # initialize the dictionary to hold the rows
  initDict(coloured_maxOfColumn, main_traffic)
  initDict(coloured_maxOfColumn, cu_traffic)
  # column loop setup
  # for each column, find max in that column highlight within some threshold
  # only wanted a few columns, but go ahead and do it for all of them anyways
  for column in main_traffic:
    # init list of blocks to color 
    columnIndex = data.columns.get_loc(column)

    maxValue = data[column].max()
    threshold = maxValue * (1 - satThreshold)
    #print("The threshold is : ", threshold, " column is : ", column)

    for row in rowsToCheck:
      num = data.iloc[row, columnIndex]

      if num >= threshold:
        coloured_maxOfColumn[column].append(row)
  #endfor

  # percent only colored if above certain threshold
  for column in cu_traffic:
    columnIndex = data.columns.get_loc(column)
    maxValue = data[column].max()/100

    # only bother if greater than .7
    if maxValue >= 0.5:
      threshold = maxValue * (1 - cuSatThreshold)

      for row in rowsToCheck:
        num = data.iloc[row, columnIndex]/100

        if num >= threshold:
          coloured_maxOfColumn[column].append(row)
      #endfor
  #endfor
  #check SIMD intensity
  #for row in rowsToCheck:
  # columnIndex = data.columns.get_loc("SIMD_MEM_Intensity")
  #  if data.iloc[row, columnIndex] > 3:
  #    coloured_maxOfColumn["SIMD_MEM_Intensity"].append(row);

# Find max in traffic columns + perfcent columns, save to maxDict
def findUniqueTiers(data, satList, tier):
  global unique_sat_node_clusters
  global unique_tiers
  #print ("Calling findUniqueTiers with : ", satList)
  # get rows that we need to check
  sat_string = ''
  if MetricName.RATE_REG_SIMD_GB_P_S in satList:
      sat_string = sat_string + 'tier_' + str(tier) + 'vr+'
  if MetricName.RATE_L1_GB_P_S in satList:
      sat_string = sat_string + 'tier_' + str(tier) + 'L1+'
  if MetricName.RATE_L2_GB_P_S in satList:
      sat_string = sat_string + 'tier_' + str(tier) + 'L2+'
  if MetricName.RATE_L3_GB_P_S in satList:
      sat_string = sat_string + 'tier_' + str(tier) + 'L3+'
  if MetricName.RATE_RAM_GB_P_S in satList:
      sat_string = sat_string + 'RAM+'

  if MetricName.RATE_FP_GFLOP_P_S in satList:
      sat_string = sat_string + 'tier_' + str(tier) + 'FLOP+'
  if MetricName.STALL_FE_PCT in satList:
      sat_string = sat_string + 'tier_' + str(tier) + 'FE+'
  if MetricName.STALL_LM_PCT in satList:
      sat_string = sat_string + 'tier_' + str(tier) + 'LM+'
  if MetricName.STALL_SB_PCT in satList:
      sat_string = sat_string + 'tier_' + str(tier) + 'SB+'
  if MetricName.STALL_RS_PCT in satList:
      sat_string = sat_string + 'tier_' + str(tier) + 'RS+'
  if MetricName.STALL_LB_PCT in satList:
      sat_string = sat_string + 'tier_' + str(tier) + 'LB+'

  if 'SIMD_MEM_Intensity' in satList:
      sat_string = sat_string + 'VEC+'

  if tier not in unique_tiers:
      unique_tiers.append(tier)

  if sat_string not in unique_sat_node_clusters:
      unique_sat_node_clusters.append(sat_string)
      target_df = pd.DataFrame(columns=data.columns.tolist())
      #target_df = target_df.append(data.iloc[row])
      #tagetFrames[sat_string] = target_df
  #print (unique_sat_node_clusters)
  #print (unique_tiers)

  for nodes in unique_sat_node_clusters:
    csv_string = nodes + '.csv'
    #tagetFrames[nodes].to_csv(csv_string, index = False, header=True)
# find unique tiers

# init target dataframe dictionary
tagetFrames = {}
# Find max in traffic columns + perfcent columns, save to maxDict
def checkCodeletTier(satdata, testCdltName, traffic, cu_traffic, sat_traffic):
  # get rows that we need to check
  codelet_in_this_tier = False
  index = satdata.index
  condition = satdata[MetricName.SHORT_NAME] == testCdltName
  test_codelet_indx = index[condition]
  row = test_codelet_indx[0]
  #satdata.to_csv(testCdltName + '_debug.csv', index = False, header=True)
  short_nameInx = satdata.columns.get_loc(MetricName.SHORT_NAME)
  short_name = satdata.iloc[test_codelet_indx[0], short_nameInx]
  for column in traffic:
    # init list of blocks to color 
    columnIndex = satdata.columns.get_loc(column)
    maxValue = satdata[column].max()
    threshold = maxValue * (1 - satThreshold)
    num = satdata.iloc[row, columnIndex]
    if num > threshold:
      codelet_in_this_tier = True
      sat_traffic.append(column)
  for column in cu_traffic:
    # init list of blocks to color 
    columnIndex = satdata.columns.get_loc(column)
    maxValue = satdata[column].max()/100
    if maxValue >= 0.5:
      threshold = maxValue * (1 - cuSatThreshold)
      num = satdata.iloc[row, columnIndex]/100
      if num > threshold:
        codelet_in_this_tier = True
        sat_traffic.append(column)
  return codelet_in_this_tier

# Find max in traffic columns + perfcent columns, save to maxDict
def findPeerCodelets(data, traffic, cu_traffic, satList, short_name):
  # get rows that we need to check
  initDict(maxOfColumn, satList)
  rowsToCheck = data.index.tolist()
  target_df = pd.DataFrame(columns=data.columns.tolist())
  csv_string = short_name + '_peer_codelet.csv'

  # To Remove the codelets that saturates nodes not in satList
  NodesNotInSatList = set(ALL_NODE_LIST) - set(satList)
  rowsToCheckForSaturation = []
  for row in rowsToCheck:
    codelet_in_non_sat_grp = False
    for column in NodesNotInSatList:
      # init list of blocks to color 
      columnIndex = data.columns.get_loc(column)
      maxValue = data[column].max()
      if column in cu_traffic:
          maxValue = maxValue/100
          if maxValue >= 0.5:
              threshold = maxValue * (1 - cuSatThreshold)
              num = data.iloc[row, columnIndex]/100
              if num > threshold:
                 codelet_in_non_sat_grp = True
      else:
          threshold = maxValue * (1 - satThreshold)
          num = data.iloc[row, columnIndex]
          if num > threshold:
            codelet_in_non_sat_grp = True
    if codelet_in_non_sat_grp is False:
      rowsToCheckForSaturation.append(row)
    #end columnfor
  #end rowfor
  #print ("The sat Threshold in findPeerCodelets :", satThreshold)
# Now find the codelets saturated.
  for row in rowsToCheckForSaturation:
    for column in satList:
      # init list of blocks to color 
      columnIndex = data.columns.get_loc(column)
      maxValue = data[column].max()
      if column in cu_traffic:
          maxValue = maxValue/100
          if maxValue >= 0.5:
              threshold = maxValue * (1 - cuSatThreshold)
              num = data.iloc[row, columnIndex]/100
              if num > threshold:
                 maxOfColumn[column].append(row)
      else:
          threshold = maxValue * (1 - satThreshold)
          num = data.iloc[row, columnIndex]
          if num > threshold:
            maxOfColumn[column].append(row)
    #end columnfor
  #end rowfor
  codelet_list = maxOfColumn[column]
  for column in satList:
      codelet_list = list(set(codelet_list) & set(maxOfColumn[column]))
  for row in codelet_list:
      target_df = target_df.append(data.iloc[row], ignore_index=True)
  #target_df.to_csv(csv_string, index = False, header=True)
  return target_df

# Find max in traffic columns + perfcent columns, save to maxDict
def createSubcluster(data, subSatList, short_name):
  # get rows that we need to check
  rowsToCheck = data.index.tolist()
  initDict(maxOfColumn, subNodeTrafficToCheck)
  target_df = pd.DataFrame(columns=data.columns.tolist())
  csv_string = short_name + '_sub_cluster.csv'
  #print (data)
  if not subSatList:
      for row in rowsToCheck:
        #print (row)
        #print ("No sub sat nodes")
        for column in subNodeTrafficToCheck:
          columnIndex = data.columns.get_loc(column)
          # init list of blocks to color 
          num = (data.iloc[row, columnIndex])/100
          if num <= 0.5:
              maxOfColumn[column].append(row);
  else:
      for row in rowsToCheck:
        #print (row)
        #print (subSatList)
        for column in subNodeTrafficToCheck and subSatList:
          # init list of blocks to color 
          columnIndex = data.columns.get_loc(column)
          num = data.iloc[row, columnIndex]
          if num >= 0.5:
              maxOfColumn[column].append(row);
    #end columnfor
  #end rowfor
  codelet_list = maxOfColumn[column]
  if not subSatList:
      for column in cuTrafficToCheck:
          codelet_list = list(set(codelet_list) & set(maxOfColumn[column]))
      for row in codelet_list:
          target_df = target_df.append(data.iloc[row], ignore_index=True)
  else :
      codelet_list = maxOfColumn[subSatList[0]]
      for column in cuTrafficToCheck and subSatList:
          codelet_list = list(set(codelet_list) & set(maxOfColumn[column]))
      for row in codelet_list:
          target_df = target_df.append(data.iloc[row], ignore_index=True)
  return target_df

# Find max in traffic columns + perfcent columns, save to maxDict
def findNextTierInColumns(data, traffic, cu_traffic):
  # get rows that we need to check
  rowsToCheck = data.index.tolist()
  target_df = pd.DataFrame(columns=data.columns.tolist())
  csv_string = 'next_tier.csv'
  for row in rowsToCheck:
    row_add_to_next_tier = True
    for column in traffic:
      # init list of blocks to color 
      columnIndex = data.columns.get_loc(column)
      maxValue = data[column].max()
      threshold = maxValue * (1 - satThreshold)
      try:
         num = data.iloc[row, columnIndex]
      except:
         print ("Cannot access row : ", row, "column : ", column)
         print ("Rows to check : ", rowsToCheck)
      if num > threshold:
        row_add_to_next_tier = False
    for column in cu_traffic:
      # init list of blocks to color
      # only bother if greater than .5
      columnIndex = data.columns.get_loc(column)
      maxValue = data[column].max()
      if maxValue >= 50:
          threshold = maxValue * (1 - cuSatThreshold)
          try:
             num = data.iloc[row, columnIndex]
          except:
             print ("Cannot access row : ", row, "column : ", column)
          if num > threshold:
             row_add_to_next_tier = False
    if row_add_to_next_tier == True:
      target_df = target_df.append(data.iloc[row], ignore_index=True)
    #end columnfor
  #end rowfor
  #target_df.to_csv(csv_string, index = False, header=True)
  return target_df

def concat_ordered_columns(frames):
    columns_ordered = []
    for frame in frames:
        columns_ordered.extend(x for x in frame.columns if x not in columns_ordered)
    final_df = pd.concat(frames, ignore_index=True)
    return final_df[columns_ordered]

def do_sub_clustering(peer_codelet_df, testDF, short_name, codelet_tier, satTrafficList):
    global result_df
    global si_passed
    global si_failed
    global no_cluster
    sub_cls_df = pd.DataFrame()
    chosen_node_set = set(BASIC_NODE_LIST)
    node_chck =  any(elem in satTrafficList  for elem in subNodeTrafficToCheck)
    if not node_chck :
        sub_nodes = []
        for column in subNodeTrafficToCheck:
          columnIndex = testDF.columns.get_loc(column)
          cu_val = testDF.iloc[0, columnIndex]/100
          if cu_val >= 0.3:
             sub_nodes.append(column)
        #print (sub_nodes)
        print ("calling createSubcluster :", short_name)
        if not sub_nodes:
           sub_node_string = "No CU Saturation in sub clustering"
        else:
           sub_node_string = str(sub_nodes) + "has 50% or more activity in sub clustering"
        sub_cluster_df = createSubcluster(peer_codelet_df, sub_nodes, short_name)
        sub_clstr_cdlt_count = sub_cluster_df.shape[0]
        if sub_clstr_cdlt_count >= 3:
            outputfile = short_name + 'SUB_Tier_'+str(codelet_tier) + "_SI.csv"
            norm = "row"
            title = "SI"
            target_df = pd.DataFrame()
            #print ("calling SI Compute with nodes :", chosen_node_set)
            sw_bias_df = compute_sw_bias(sub_cluster_df)
            compute_and_plot('XFORM', sub_cluster_df, outputfile, norm, title, chosen_node_set, target_df)
            peer_dfs = [sub_cluster_df,testDF]
            final_df = concat_ordered_columns(peer_dfs)
            result = test_and_plot_orig('ORIG', final_df, outputfile, norm, title, chosen_node_set, target_df, short_name)
            if DO_DEBUG_LOGS:
               final_df.to_csv(short_name+'_sub_cluster_report.csv', index = True, header=True)
            #print ("Sub_Cluster Result is : ", result)
            if result == True :
                print (short_name, "Passed the Sub_Cluster SI Test =>")
                si_passed +=1
                findUniqueTiers(sub_cluster_df, satTrafficList, codelet_tier)
                #result_df = result_df.append({'ShortName' : short_name, 'peer_codelet_cnt' : sub_clstr_cdlt_count,
                #            'Tier' : 'SUB_Tier_'+str(codelet_tier), 'Sat_Node' : satTrafficList, 'Sat_Sub_Node' : sub_node_string, 'SI_Result' : 'Outside Box',
                #            'GFlops' : final_df.loc[final_df[MetricName.SHORT_NAME] == short_name, MetricName.RATE_FP_GFLOP_P_S].item(),
                #            'Saturation' : final_df.loc[final_df[MetricName.SHORT_NAME] == short_name, 'Saturation'].item(),
                #            'Intensity' : final_df.loc[final_df[MetricName.SHORT_NAME] == short_name, 'Intensity'].item()}, 
                #ignore_index = True)
                sub_cls_df = pd.DataFrame({MetricName.SHORT_NAME : short_name, 'peer_codelet_cnt' : sub_clstr_cdlt_count,
                            'Tier' : 'SUB_Tier_'+str(codelet_tier), 'Sat_Node' : satTrafficList, 'Sat_Sub_Node' : sub_node_string, 'SI_Result' : 'Outside Box',
                            'GFlops' : final_df.loc[final_df[MetricName.SHORT_NAME] == short_name, MetricName.RATE_FP_GFLOP_P_S].item(),
                            'Saturation' : final_df.loc[final_df[MetricName.SHORT_NAME] == short_name, 'Saturation'].item(),
                            'Intensity' : final_df.loc[final_df[MetricName.SHORT_NAME] == short_name, 'Intensity'].item()})
            else:
                print (short_name, "Failed the Sub_ClusterSI Test =>")
                si_failed +=1
                findUniqueTiers(sub_cluster_df, satTrafficList, codelet_tier)
                #result_df = result_df.append({'ShortName' : short_name, 'peer_codelet_cnt' : sub_clstr_cdlt_count,
                #            'Tier' : 'SUB_Tier_'+str(codelet_tier), 'Sat_Node' : satTrafficList, 'Sat_Sub_Node' : sub_node_string, 'SI_Result' : 'Inside Box',
                #            'GFlops' : final_df.loc[final_df[MetricName.SHORT_NAME] == short_name, MetricName.RATE_FP_GFLOP_P_S].item(),
                #            'Saturation' : final_df.loc[final_df[MetricName.SHORT_NAME] == short_name, 'Saturation'].item(),
                #            'Intensity' : final_df.loc[final_df[MetricName.SHORT_NAME] == short_name, 'Intensity'].item()},
                #ignore_index = True)
                sub_cls_df = pd.DataFrame({MetricName.SHORT_NAME : short_name, 'peer_codelet_cnt' : sub_clstr_cdlt_count,
                            'Tier' : 'SUB_Tier_'+str(codelet_tier), 'Sat_Node' : satTrafficList, 'Sat_Sub_Node' : sub_node_string, 'SI_Result' : 'Inside Box',
                            'GFlops' : final_df.loc[final_df[MetricName.SHORT_NAME] == short_name, MetricName.RATE_FP_GFLOP_P_S].item(),
                            'Saturation' : final_df.loc[final_df[MetricName.SHORT_NAME] == short_name, 'Saturation'].item(),
                            'Intensity' : final_df.loc[final_df[MetricName.SHORT_NAME] == short_name, 'Intensity'].item()})
        else:
            print (short_name, "Not enough codelets in the Sub_ClusterSI Test =>")
            no_cluster+=1
            findUniqueTiers(sub_cluster_df, satTrafficList, codelet_tier)
            #result_df = result_df.append({'ShortName' : short_name, 'peer_codelet_cnt' : sub_clstr_cdlt_count,
            #           'Tier' : 'SUB_Tier_'+str(codelet_tier), 'Sat_Node' : satTrafficList, 'Sat_Sub_Node' : sub_node_string, 'SI_Result' : 'No Sub Cluster'})
            sub_cls_df = pd.DataFrame({MetricName.SHORT_NAME : short_name, 'peer_codelet_cnt' : sub_clstr_cdlt_count,
                        'Tier' : 'SUB_Tier_'+str(codelet_tier), 'Sat_Node' : satTrafficList, 'Sat_Sub_Node' : sub_node_string, 'SI_Result' : 'No Sub Cluster'})
    else :
        print (short_name, "How to do Sub_Cluster SI Test ??=>")
        sub_cls_df = pd.DataFrame({MetricName.SHORT_NAME : short_name, 'peer_codelet_cnt' : 0,
                   'Tier' : 'SUB_Tier_'+str(codelet_tier), 'Sat_Node' : satTrafficList, 'Sat_Sub_Node' : 'N/A', 'SI_Result' : 'No Sub Cluster'})
        si_failed+=1
    return sub_cls_df

def find_swbias_cluster(satSetDF, testDF, swbias_threshold):
    mask = satSetDF[NonMetricName.SI_SW_BIAS] >= swbias_threshold
    if testDF.iloc[0]['Net_SW_Bias'] >= swbias_threshold:
         return satSetDF[mask]
    else:
         return satSetDF[~mask]

def do_swbias_clustering(peer_codelet_df, testDF, satTrafficList):
    sw_bias_df = compute_sw_bias(peer_codelet_df)
    peer_codelet_swBias_df = pd.merge(peer_codelet_df, sw_bias_df)
    sw_bias_df = compute_sw_bias(testDF)
     # get SW bias Vector
    neg_bias_vec =[sw_bias_df['Nd_CNVT_OPS'].item(), sw_bias_df['Nd_DIV_OPS'].item(), sw_bias_df['Nd_clu_score'].item(),
                   sw_bias_df['Nd_Recurrence'].item(), sw_bias_df['Nd_RHS'].item()] 
    pos_bias_vec =[sw_bias_df['Nd_VEC_OPS'].item(), sw_bias_df['Nd_ISA_EXT_TYPE'].item(), sw_bias_df['Nd_FMA_OPS'].item()]

    test_swBias_df = pd.merge(testDF, sw_bias_df)
    swbias_cluster_df = find_swbias_cluster(peer_codelet_swBias_df, test_swBias_df, 0)
    cdlt_count = swbias_cluster_df.shape[0]
    chosen_node_set = set(BASIC_NODE_LIST)
    short_name = str(testDF.iloc[0][MetricName.SHORT_NAME])
    if cdlt_count >= 3:
        for elem in satTrafficList:
            if elem in CU_NODE_SET:
                #print (CU_NODE_DICT[elem])
                chosen_node_set.add(CU_NODE_DICT[elem])
        outputfile = short_name + '_sw_bias_'+ "_SI.csv"
        norm = "row"
        title = "SI"
        target_df = pd.DataFrame()
        compute_and_plot('XFORM', swbias_cluster_df, outputfile, norm, title, chosen_node_set, target_df)
        peer_dfs = [swbias_cluster_df,test_swBias_df]
        final_df = concat_ordered_columns(peer_dfs)
        final_df.to_csv(outputfile, index = False, header=True)
        bias_result = test_and_plot_orig('ORIG', final_df, outputfile, norm, title, chosen_node_set, target_df, short_name)
        s_length = swbias_cluster_df['Saturation'].max() - swbias_cluster_df['Saturation'].min()
        i_length = swbias_cluster_df['Intensity'].max() - swbias_cluster_df['Intensity'].min()
        box_length = '{' + str(round(s_length, 2)) + ' , ' + str(round(i_length, 2)) + '}'

        s_ratio = swbias_cluster_df['Saturation'].max() / swbias_cluster_df['Saturation'].min()
        i_ratio = swbias_cluster_df['Intensity'].max() / swbias_cluster_df['Intensity'].min()
        box_ratio = '{' + str(round(s_length, 2)) + ' , ' + str(round(i_length, 2)) + '}'

        if bias_result == True :
            print (short_name, " swbias cluster Outside the SI Box =>")
            result_df = pd.DataFrame({MetricName.SHORT_NAME : short_name, 'peer_codelet_cnt' : cdlt_count,'SI_Result' : 'Outside Box',
                        'Box_Length' : box_length, 'Box_Ratio' : box_ratio, 'Neg_SW_bias_Vec' : str(neg_bias_vec), 'Pos_SW_bias_Vec' : str(pos_bias_vec)}, index=[0])
        else:
            print (short_name, " swbias cluster Inside the SI Box =>")
            result_df = pd.DataFrame({MetricName.SHORT_NAME : short_name, 'peer_codelet_cnt' : cdlt_count,'SI_Result' : 'Inside Box',
                        'Box_Length' : box_length, 'Box_Ratio' : box_ratio, 'Neg_SW_bias_Vec' : str(neg_bias_vec), 'Pos_SW_bias_Vec' : str(pos_bias_vec)}, index=[0])
    else:
        print (short_name, " swbias cluster not present =>")
        result_df = pd.DataFrame({MetricName.SHORT_NAME : short_name, 'peer_codelet_cnt' : 0,'SI_Result' : 'No Cluster',
        'Box_Length' : 0, 'Box_Ratio' : 0, 'Neg_SW_bias_Vec' : str(neg_bias_vec), 'Pos_SW_bias_Vec' : str(pos_bias_vec)}, index=[0])
    return result_df
def find_cluster(satSetDF, testDF, short_name, codelet_tier):
    global result_df
    global si_passed
    global si_failed
    global no_cluster
    global satThreshold
    peer_cdlt_count = 0
    dfs = [satSetDF,testDF]
    full_df = concat_ordered_columns(dfs)
    #full_df.to_csv(short_name + '_debug.csv', index = False, header=True)

    satTrafficList = []
    codelet_tier = codelet_tier + 1;
    #satThreshold = satThreshold*codelet_tier
    satThreshold = 0.1*codelet_tier
    max_mem_traffic = testDF[[MetricName.RATE_L1_GB_P_S, MetricName.RATE_L2_GB_P_S, MetricName.RATE_L3_GB_P_S, MetricName.RATE_RAM_GB_P_S]].max(axis=1)
    mem_traffic_threshold = 0.1 * max_mem_traffic.item()
    tstcdlt_TrafficToCheck = []
    tstcdlt_TrafficToCheck.append(MetricName.RATE_FP_GFLOP_P_S)
    for column in memTrafficToCheck:
       columnIndex = testDF.columns.get_loc(column)
       val = testDF.iloc[0, columnIndex]
       if (val > mem_traffic_threshold):
          tstcdlt_TrafficToCheck.append(column)
    check_codlet_in_this_tier = checkCodeletTier(full_df, short_name, tstcdlt_TrafficToCheck, percentsToCheck, satTrafficList)

    # Print the tiering process : For debug only
    if PRINT_COLOURED_TIERS:
       tier_book_path = short_name + "_tier.xlsx"
       tier_book = load_workbook(tier_book_path);
       highlightSheet = tier_book.create_sheet(short_name + "_" + str(codelet_tier), 0)
       sheet_title = short_name + "_" + str(codelet_tier) + "_" + "Highlights"
       # import all rows into xlsx package
       for r in dataframe_to_rows(full_df, index=False, header=True):
          highlightSheet.append(r)
       findMaxInColumnsToColor(full_df, trafficToCheck, percentsToCheck)
       colorMaxInColumn(full_df, coloured_maxOfColumn, opBlue, highlightSheet)
       tier_book.save(tier_book_path)

    if check_codlet_in_this_tier == True:
        #satSetDF.to_csv(short_name+'_tier_report.csv', index = True, header=True)
        peer_codelet_df = findPeerCodelets(satSetDF, trafficToCheck, percentsToCheck, satTrafficList, short_name)
        peer_cdlt_count = peer_codelet_df.shape[0]
        chosen_node_set = set(BASIC_NODE_LIST)
        sat_rng_string = ''
        sat_node_string = ''
        for i, elem in enumerate(satTrafficList):
            if i:
                sat_rng_string += "], "
            sat_rng_string += elem
            sat_rng_string += ": ["
            sat_rng_string += str(peer_codelet_df[elem].max())
            sat_rng_string += "  "
            sat_rng_string += str(peer_codelet_df[elem].min())
            sat_rng_string += "] "

        for i, elem in enumerate(satTrafficList):
            if i:
               sat_node_string += " , "
            sat_node_string += elem

        if peer_cdlt_count >= 3:
            for elem in satTrafficList:
                if elem in CU_NODE_SET:
                    #print (CU_NODE_DICT[elem])
                    chosen_node_set.add(CU_NODE_DICT[elem])
            outputfile = short_name + 'Tier_'+str(codelet_tier) + "_SI.csv"
            norm = "row"
            title = "SI"
            target_df = pd.DataFrame()
            #print ("calling SI Compute with nodes :", chosen_node_set)
            compute_and_plot('XFORM', peer_codelet_df, outputfile, norm, title, chosen_node_set, target_df)
            peer_dfs = [peer_codelet_df,testDF]
            final_df = concat_ordered_columns(peer_dfs)
            sw_bias_df = compute_sw_bias(final_df)
            final_df = pd.merge(final_df, sw_bias_df)
            result = test_and_plot_orig('ORIG', final_df, outputfile, norm, title, chosen_node_set, target_df, short_name)

            #print ("Saturation of : ", short_name, " : ", final_df.loc[final_df[MetricName.SHORT_NAME] == short_name, 'Saturation'].item())
 
            s_length = peer_codelet_df['Saturation'].max() - peer_codelet_df['Saturation'].min()
            i_length = peer_codelet_df['Intensity'].max() - peer_codelet_df['Intensity'].min()
            box_length = '{' + str(round(s_length, 2)) + ' , ' + str(round(i_length, 2)) + '}'
            peer_codelet_df['Norm_Tier'] = codelet_tier + ((peer_codelet_df['Saturation'] - peer_codelet_df['Saturation'].min())/s_length)
            print(peer_codelet_df['Norm_Tier'])

            s_ratio = peer_codelet_df['Saturation'].max() / peer_codelet_df['Saturation'].min()
            i_ratio = peer_codelet_df['Intensity'].max() / peer_codelet_df['Intensity'].min()
            box_ratio = '{' + str(round(s_ratio, 2)) + ' , ' + str(round(i_ratio, 2)) + '}'
            # Do SW_BIAS Clustering anyways
            bias_res_df = do_swbias_clustering(peer_codelet_df,testDF, satTrafficList)
            #if (s_length > 1.5 or i_length > 1.5):
            #    bias_res_df = do_swbias_clustering(peer_codelet_df,testDF, satTrafficList)
            codelet_variant = testDF['Variant'].item()
            codelet_set = testDF['Set'].item()
            if result == True :
                print (short_name, "Passed the SI Test =>")
                si_passed +=1
                findUniqueTiers(peer_codelet_df, satTrafficList, codelet_tier)
                result_df = result_df.append({'ShortName' : short_name, 'peer_codelet_cnt' : peer_cdlt_count,
                   'Tier' : str(codelet_tier), 'Sat_Node' : sat_node_string, 'Sat_Range' : sat_rng_string, 'SI_Result' : 'Outside Box',
                   'Box_Length' : box_length,
                   'Box_Ratio' : box_ratio,
                   'Variant' : codelet_variant,
                   'Set' : codelet_set,
                   'GFlops' : round(final_df.loc[final_df[MetricName.SHORT_NAME] == short_name, MetricName.RATE_FP_GFLOP_P_S].item(), 2),
                   'Saturation' : round(final_df.loc[final_df[MetricName.SHORT_NAME] == short_name, 'Saturation'].item(), 2),
                   'Intensity' : round(final_df.loc[final_df[MetricName.SHORT_NAME] == short_name, 'Intensity'].item(), 2),
                   'SW_bias' : round(final_df.loc[final_df[MetricName.SHORT_NAME] == short_name, 'Net_SW_Bias'].item(), 2),
                   'neg_bias[CNVT,DIV,clu_score,Rec,RHS]' : bias_res_df.loc[bias_res_df[MetricName.SHORT_NAME] == short_name, 'Neg_SW_bias_Vec'].item(),
                   'pos_bias[VEC_OPS,ISA_EXT,FMA_OPS]' : bias_res_df.loc[bias_res_df[MetricName.SHORT_NAME] == short_name, 'Pos_SW_bias_Vec'].item(),
                   'SW_bias_Result' : bias_res_df.loc[bias_res_df[MetricName.SHORT_NAME] == short_name, 'SI_Result'].item(),
                   'SW_bias_box_length' : bias_res_df.loc[bias_res_df[MetricName.SHORT_NAME] == short_name, 'Box_Length'].item(),
                   'SW_bias_box_ratio' : bias_res_df.loc[bias_res_df[MetricName.SHORT_NAME] == short_name, 'Box_Ratio'].item(),
                   'SW_bias_CLS_CDLTS' : bias_res_df.loc[bias_res_df[MetricName.SHORT_NAME] == short_name, 'peer_codelet_cnt'].item()},
                ignore_index = True) 
                if DO_SUB_CLUSTERING:
                   sub_df = do_sub_clustering(peer_codelet_df, testDF, short_name, codelet_tier, satTrafficList)
                   print ("The sub_node saturated : ", sub_df['Sat_Sub_Node'])
            else:
                print (short_name, "Failed the SI Test =>")
                si_failed +=1
                if DO_SUB_CLUSTERING:
                   sub_df = do_sub_clustering(peer_codelet_df, testDF, short_name, codelet_tier, satTrafficList)
                   print ("The sub_node saturated : ", sub_df['Sat_Sub_Node'])
                result_df = result_df.append({'ShortName' : short_name, 'peer_codelet_cnt' : peer_cdlt_count,
                            'Tier' : str(codelet_tier), 'Sat_Node' : sat_node_string, 'Sat_Range' : sat_rng_string, 'SI_Result' : 'Inside Box',
                            'Box_Length' : box_length,
                            'Box_Ratio' : box_ratio,
                            'Variant' : codelet_variant,
                            'Set' : codelet_set,
                            'GFlops' : round(final_df.loc[final_df[MetricName.SHORT_NAME] == short_name, MetricName.RATE_FP_GFLOP_P_S].item(), 2),
                            'Saturation' : round(final_df.loc[final_df[MetricName.SHORT_NAME] == short_name, 'Saturation'].item(), 2),
                            'Intensity' : round(final_df.loc[final_df[MetricName.SHORT_NAME] == short_name, 'Intensity'].item(), 2),
                            'SW_bias' : round(final_df.loc[final_df[MetricName.SHORT_NAME] == short_name, 'Net_SW_Bias'].item(), 2),
                            'neg_bias[CNVT,DIV,clu_score,Rec,RHS]' : bias_res_df.loc[bias_res_df[MetricName.SHORT_NAME] == short_name, 'Neg_SW_bias_Vec'].item(),
                            'pos_bias[VEC_OPS,ISA_EXT,FMA_OPS]' : bias_res_df.loc[bias_res_df[MetricName.SHORT_NAME] == short_name, 'Pos_SW_bias_Vec'].item(),
                            'SW_bias_Result' : bias_res_df.loc[bias_res_df[MetricName.SHORT_NAME] == short_name, 'SI_Result'].item(),
                            'SW_bias_box_length' : bias_res_df.loc[bias_res_df[MetricName.SHORT_NAME] == short_name, 'Box_Length'].item(),
                            'SW_bias_box_ratio' : bias_res_df.loc[bias_res_df[MetricName.SHORT_NAME] == short_name, 'Box_Ratio'].item(),
                            'SW_bias_CLS_CDLTS' : bias_res_df.loc[bias_res_df[MetricName.SHORT_NAME] == short_name, 'peer_codelet_cnt'].item()},
                ignore_index = True)

            if DO_DEBUG_LOGS:
                final_df.to_csv(short_name+'_report.csv', index = True, header=True)
        else:
            print (short_name, "No Cluster for the SI Test =>")
            sw_bias_df = compute_sw_bias(testDF)
            # get SW bias Vector
            neg_bias_vec =[sw_bias_df['Nd_CNVT_OPS'].item(), sw_bias_df['Nd_DIV_OPS'].item(), sw_bias_df['Nd_clu_score'].item(),
                           sw_bias_df['Nd_Recurrence'].item(), sw_bias_df['Nd_RHS'].item()] 
            pos_bias_vec =[sw_bias_df['Nd_VEC_OPS'].item(), sw_bias_df['Nd_ISA_EXT_TYPE'].item(), sw_bias_df['Nd_FMA_OPS'].item()]
            no_cluster+=1
            findUniqueTiers(peer_codelet_df, satTrafficList, codelet_tier)
            result_df = result_df.append({'ShortName' : short_name, 'peer_codelet_cnt' : peer_cdlt_count,
                        'Tier' : str(codelet_tier), 'Sat_Node' : sat_node_string, 'Sat_Range' : sat_rng_string, 'SI_Result' : 'No Cluster',
                        'GFlops' : round(testDF.loc[testDF[MetricName.SHORT_NAME] == short_name, MetricName.RATE_FP_GFLOP_P_S].item(), 2),
                        'SW_bias' : round(testDF.loc[testDF[MetricName.SHORT_NAME] == short_name, 'Net_SW_Bias'].item(), 2),
                        'neg_bias[CNVT,DIV,clu_score,Rec,RHS]' : neg_bias_vec,
                        'pos_bias[VEC_OPS,ISA_EXT,FMA_OPS]' : pos_bias_vec}, 
                        #testDF.loc[testDF[MetricName.SHORT_NAME] == short_name, 'Net_SW_Bias'].item()},
            ignore_index = True) 
    else:
        next_tier_df = findNextTierInColumns(satSetDF, trafficToCheck, percentsToCheck)
        #print ("next tier codelet count : ", next_tier_df.shape[0])
        if next_tier_df.shape[0] > 5 :
            find_cluster(next_tier_df, testDF, short_name, codelet_tier)
        else :
                print (short_name, "Last Tier: No Cluster for the SI Test =>")
                no_cluster+=1
                result_df = result_df.append({'ShortName' : short_name, 'peer_codelet_cnt' : peer_cdlt_count,
                            'Tier' : 'LastTier_'+str(codelet_tier), 'Sat_Node' : sat_node_string, 'SI_Result' : 'No Cluster',
                            'GFlops' : round(testDF.loc[testDF[MetricName.SHORT_NAME] == short_name, MetricName.RATE_FP_GFLOP_P_S].item(), 2),
                            'SW_bias' : 0 }, #testDF.loc[testDF[MetricName.SHORT_NAME] == short_name, 'Net_SW_Bias'].item()},
                ignore_index = True) 

def find_all_clusters(satSetDF):
    global satThreshold;
    print ("Memory Node Saturation Threshold : ", satThreshold)
    print ("Control Node Saturation Threshold : ", cuSatThreshold)

def do_sat_analysis(satSetDF,testSetDF):
    short_name=''
    # Creating an empty Dataframe with column names only
    #print("Empty Dataframe ", dfObj, sep='\n')
    global satThreshold;
    codelet_tested = 0;
    print ("Memory Node Saturation Threshold : ", satThreshold)
    print ("Control Node Saturation Threshold : ", cuSatThreshold)
    for i, row in testSetDF.iterrows():
        satThreshold = 0.1;
        l_df = satSetDF
        testDF = pd.DataFrame(columns=testSetDF.columns.tolist())
        short_name = str(row[MetricName.SHORT_NAME])
        variant = str(row[MetricName.VARIANT])
        short_name = 'test-' + short_name + '-' + variant
        filename=short_name[-19:]
        re.sub('[^\w\-_\. ]', '_', filename)
        row[MetricName.SHORT_NAME] = short_name
        testDF = testDF.append(row, ignore_index=False)[testSetDF.columns.tolist()]
        codelet_tested += 1;
        print ("Sat Analysis on ", short_name, ": ",  satThreshold)
        #if RUN_SW_BIAS:
            #sw_bias_df = compute_sw_bias(testDF)
            #resul_df = pd.merge(testDF, sw_bias_df)
        #    if DO_DEBUG_LOGS:
                #resul_df.to_csv(filename + '_sw_bias.csv', index = True, header=True)
        find_cluster(satSetDF, testDF, short_name, 0)
        #print ("codelet_tested = ", codelet_tested, " Passed = ", si_passed, " Failed = ", si_failed, " No Cluster = ", no_cluster)
    result_df.to_csv('Result_report.csv', index = True, header=True)
    print ("Total No. of codelets tested : ", codelet_tested)
    print ("Total No. of codelets outside SI box: ", si_passed)
    print ("Total No. of codelets inside SI box: ", si_failed)
    print ("Total No. of codelets without Cluster : ", no_cluster)
    print ("Total No. of Tiers : ", len(unique_tiers))
    print ("Total No. of unique clusters : ", len(unique_sat_node_clusters))
    stat_data = [{'Mem Threshold' : satThreshold, 'CU Threshold' : cuSatThreshold,
            'Codelets Tested' : codelet_tested, 'Passed' : si_passed, 'Failed' : si_failed,
            'No Cluster' : no_cluster, 'Tier Count' : len(unique_tiers), 'Cluster Count' : len(unique_sat_node_clusters)}]
    stats_df = pd.DataFrame(stat_data)
    my_file = Path("./statistcs.csv")
    if my_file.is_file():
       # file exists
       stats_df.to_csv('statistcs.csv', mode='a', header=False, index=False)
    else :
       stats_df.to_csv('statistcs.csv', mode='w', header=True, index=False)


def main(argv):
    inputfile = []
    global satThreshold
    global cuSatThreshold
    sys.setrecursionlimit(10**9) 
  # if 3 arg specified, assumes 3rd is threshold replacement
    #print("No of sys args : ", len(sys.argv))
    if (len(sys.argv) >= 5):
      satThreshold = float(sys.argv[3])
      cuSatThreshold = float(sys.argv[4])

    print("Attempting to read", csvToRead)

    # read into pandas
    mainDataFrame = pd.read_csv(csvToRead)
    TestSetDF = pd.read_csv(csvTestSet)
    grouped = TestSetDF.groupby(MetricName.VARIANT)
    mask = (TestSetDF['Set'] == 'BENEFITTING') & (TestSetDF['Variant'] == 'ORIG')
    #mask = (TestSetDF['Set'] == 'BENEFITTING')

    print("Read successful!")

    # save original dataframe
    originalDataFrame = mainDataFrame

    # SIMD_RATE divided by max traffic column for memory
    addAfterColumn = mainDataFrame.columns.get_loc(MetricName.RATE_RAM_GB_P_S) + 1
    mainDataFrame.insert(addAfterColumn, "SIMD_MEM_Intensity",
      mainDataFrame[MetricName.RATE_REG_SIMD_GB_P_S] / mainDataFrame[[MetricName.RATE_L1_GB_P_S, MetricName.RATE_L2_GB_P_S, MetricName.RATE_L3_GB_P_S, MetricName.RATE_RAM_GB_P_S]].max(axis=1))

    # FLOP_RATE divided by max traffic column for memory
    addAfterColumn = mainDataFrame.columns.get_loc(MetricName.RATE_FP_GFLOP_P_S) + 1
    mainDataFrame.insert(addAfterColumn, "FLOP_MEM_Intensity",
    mainDataFrame[MetricName.RATE_FP_GFLOP_P_S] / mainDataFrame[[MetricName.RATE_L1_GB_P_S, MetricName.RATE_L2_GB_P_S, MetricName.RATE_L3_GB_P_S, MetricName.RATE_RAM_GB_P_S]].max(axis=1)/8)
    if RUN_SI:
      #do_sat_analysis(mainDataFrame, TestSetDF[mask])
      do_sat_analysis(mainDataFrame, TestSetDF)
    if PRINT_ALL_CLUSTERS:
      find_all_clusters(mainDataFrame)


if __name__ == "__main__":
    main(sys.argv[1:])
